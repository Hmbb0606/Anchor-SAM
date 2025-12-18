import os
from pathlib import Path
import time
import glob
import numpy as np
import torch.nn.functional as F
import torch
import logging
from tqdm import tqdm
from colorama import Fore, Style, init
import inspect
import openpyxl

init(autoreset=True)
EXCEL_LOG_FILE = 'training_metrics.xlsx'


def de_normalize(tensor, mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]):

    tensor_copy = tensor.clone().cpu()
    mean = torch.tensor(mean, device=tensor_copy.device).view(3, 1, 1)
    std = torch.tensor(std, device=tensor_copy.device).view(3, 1, 1)
    tensor_copy.mul_(std).add_(mean)
    tensor_copy = torch.clamp(tensor_copy, 0, 1)
    return tensor_copy

def save_model(model, path, epoch, mode, optimizer=None, scheduler=None, total_step=None, best_metrics=None):

    Path(path).mkdir(parents=True, exist_ok=True)

    if mode in ['f1score', 'loss', 'test_f1score']:
        old_models = glob.glob(os.path.join(path, f'best_{mode}_*.pth'))
        for old_model in old_models:
            try:
                os.remove(old_model)
                logging.info(f"Removed old best model: {old_model}")
            except OSError as e:
                logging.error(f"Error removing file {old_model}: {e}")

    # 保存检查点
    if mode == 'checkpoint':
        state_dict = {'net': model.state_dict(),
                      'optimizer': optimizer.state_dict(),
                      'scheduler': scheduler.state_dict(),
                      'total_step': total_step,
                      'epoch': epoch,
                      'best_metrics': best_metrics
                      }
        filename = f'checkpoint_{epoch}.pth'
        torch.save(state_dict, str(os.path.join(path, filename)))
    else:  # 保存最佳模型时，只保存模型权重
        filename = f'best_{mode}_{epoch}.pth'
        torch.save(model.state_dict(), str(os.path.join(path, filename)))

    if mode in ['f1score', 'loss', 'test_f1score']:
        logging.info(f'New best {mode} model saved as "{filename}" at epoch {epoch}!')
    else:
        logging.info(f'Checkpoint saved as "{filename}" at epoch {epoch}!')

# 图像切块 (无变化)
def process_image_patches(image, labels, crop_size, mode='train'):
    """图像切块处理。"""
    b, c, h, w = image.shape
    if mode == 'train' or mode == 'val' or mode == 'test':
        num_patches_h = h // crop_size
        num_patches_w = w // crop_size
        used_h = num_patches_h * crop_size
        used_w = num_patches_w * crop_size
        image_cropped = image[:, :, :used_h, :used_w]
        labels_cropped = labels[:, :used_h, :used_w]
        image_patches = image_cropped.unfold(2, crop_size, crop_size).unfold(3, crop_size, crop_size)
        B, C, new_H, new_W, _, _ = image_patches.size()
        image = image_patches.permute(0, 2, 3, 1, 4, 5).reshape(-1, C, crop_size, crop_size).contiguous()
        labels_patches = labels_cropped.unfold(1, crop_size, crop_size).unfold(2, crop_size, crop_size)
        labels = labels_patches.reshape(-1, crop_size, crop_size).contiguous()
        return image, labels
    else:
        pad_h = (crop_size - h % crop_size) % crop_size
        pad_w = (crop_size - w % crop_size) % crop_size
        if pad_h > 0 or pad_w > 0:
            image = F.pad(image, (0, pad_w, 0, pad_h), mode='constant', value=1.0)
            labels = F.pad(labels.unsqueeze(1), (0, pad_w, 0, pad_h), mode='constant', value=0.0).squeeze(1)
        image_patches = image.unfold(2, crop_size, crop_size).unfold(3, crop_size, crop_size)
        B, C, new_H, new_W, _, _ = image_patches.size()
        image = image_patches.permute(0, 2, 3, 1, 4, 5).reshape(-1, C, crop_size, crop_size).contiguous()
        labels_patches = labels.unfold(1, crop_size, crop_size).unfold(2, crop_size, crop_size)
        labels = labels_patches.reshape(-1, crop_size, crop_size).contiguous()
        return image, labels

# 记录 Batch 指标到 Comet
def _log_batch_metrics_to_comet(experiment, metrics, current_lr, mode, loss, diceloss, foclaloss, total_step):
    """将单个批次的指标记录到Comet。"""
    metrics_to_log = {
        f'batch_{mode}_loss': loss.item(),
        f'batch_{mode}_accuracy': metrics['accuracy'].item(),
        f'batch_{mode}_precision': metrics['precision'].item(),
        f'batch_{mode}_recall': metrics['recall'].item(),
        f'batch_{mode}_f1score': metrics['f1score'].item(),
        f'batch_{mode}_iou': metrics['iou'].item(),
        'learning_rate': current_lr,
        f'batch_{mode}_loss_dice': diceloss.item(),
        f'batch_{mode}_loss_focal': foclaloss.item(),
    }
    experiment.log_metrics(metrics_to_log, step=total_step)

# 记录 Epoch 总结到 Comet
def _log_epoch_summary_to_comet(experiment, epoch_metrics, sample_images, epoch, mode, sample_info={}):
    """将一个轮次的总结指标和样本图像记录到Comet。"""
    metrics_to_log = {f'epoch_{mode}_{k}': v for k, v in epoch_metrics.items()}
    experiment.log_metrics(metrics_to_log, step=epoch)

    sample_image, sample_label, sample_pred = sample_images
    if sample_image is not None:
        temp_image_dir = './comet_logs/images_log/'
        os.makedirs(temp_image_dir, exist_ok=True)
        large_name = sample_info.get("name", "unknown")
        patch_idx = sample_info.get("patch_idx", "unknown")
        base_name = f"{mode}_epoch{epoch}_{large_name}_patch_{patch_idx}".replace('/', '_').replace('\\', '_')
        input_img_path = os.path.join(temp_image_dir, f'{base_name}_input.png')
        truth_img_path = os.path.join(temp_image_dir, f'{base_name}_truth.png')
        pred_img_path = os.path.join(temp_image_dir, f'{base_name}_pred.png')
        try:
            sample_image.save(input_img_path)
            sample_label.save(truth_img_path)
            sample_pred.save(pred_img_path)
            experiment.log_image(input_img_path, name=f'{base_name}_input.png', step=epoch)
            experiment.log_image(truth_img_path, name=f'{base_name}_truth.png', step=epoch)
            experiment.log_image(pred_img_path, name=f'{base_name}_pred.png', step=epoch)
        except Exception as e:
            print(f"\n{Fore.RED}[ERROR] Failed to save or log image to Comet: {e}{Style.RESET_ALL}")


# 处理验证逻辑
def _handle_best_model_saving(net, optimizer, scheduler, total_step, epoch_metrics, epoch, best_metrics, config, paths,
                              mode):
    """根据模式(val/test)处理最佳模型的保存逻辑"""
    checkpoint_path, best_f1_path, best_loss_path = paths

    if mode == 'val':
        current_f1 = epoch_metrics.get('f1score', 0.0)
        current_loss = epoch_metrics.get('loss_total', float('inf'))

        if current_f1 > best_metrics['best_f1score']:
            best_metrics['best_f1score'] = current_f1
            if config.save_best_model:
                save_model(net, best_f1_path, epoch, 'f1score')

        if current_loss < best_metrics['lowest loss']:
            best_metrics['lowest loss'] = current_loss
            if config.save_best_model:
                save_model(net, best_loss_path, epoch, 'loss')

        if config.save_checkpoint and checkpoint_path is not None and (epoch + 1) % config.save_interval == 0:
            save_model(net,
                       checkpoint_path,
                       epoch,
                       'checkpoint',
                       optimizer=optimizer,
                       scheduler=scheduler,
                       total_step=total_step,
                       best_metrics=best_metrics
                       )

    elif mode == 'test':
        current_f1 = epoch_metrics.get('f1score', 0.0)
        if current_f1 > best_metrics['best_test_f1score']:
            best_metrics['best_test_f1score'] = current_f1
            if config.save_best_model:
                save_model(net, best_f1_path, epoch, 'test_f1score')

    return best_metrics

# 打印Epoch总结
def print_epoch_summary(epoch, mode, epoch_metrics):
    """格式化打印一个Epoch的总结。"""
    print(f"\n{'=' * 50}")
    print(f"EPOCH {epoch} {mode.upper()} SUMMARY")
    print(f"{'=' * 50}")
    for k, v in epoch_metrics.items():
        print(f"{k.capitalize():<15}: {v:.4f}")
    print(f"{'=' * 50}\n")



# 记录 Epoch 指标到 Excel
def log_epoch_metrics_to_excel(epoch, mode, epoch_metrics):
    """
    将指标记录到Excel。此版本修正了第一行为空的问题，
    确保表头在第一行，数据从第二行开始。
    """
    sheet_name = mode
    data_row = [epoch] + list(epoch_metrics.values())

    try:

        if os.path.exists(EXCEL_LOG_FILE):
            workbook = openpyxl.load_workbook(EXCEL_LOG_FILE)
        else:
            workbook = openpyxl.Workbook()
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']

            headers = ['Epoch'] + list(epoch_metrics.keys())

            for sn in ['train', 'val', 'test']:
                sheet = workbook.create_sheet(sn)
                sheet.append(headers)

        sheet = workbook[sheet_name]

        sheet.append(data_row)

        workbook.save(EXCEL_LOG_FILE)

    except Exception as e:
        print(f"\n{Fore.RED}[ERROR] Failed to write metrics to Excel file {EXCEL_LOG_FILE}: {e}{Style.RESET_ALL}")


def train_val_test(
        mode, dataset_name,
        dataloader, device, experiment, net, optimizer,
        total_step, lr, criterion, metric_collection, to_pilimg, epoch,
        config, scheduler=None,
        grad_scaler=None,
        best_metrics=None, checkpoint_path=None,
        best_f1score_model_path=None, best_loss_model_path=None, forward_kwargs=None
):

    net.train() if mode == 'train' else net.eval()
    logging.info(f'SET model mode to {mode}!')

    epoch_loss = 0
    metric_collection.reset()
    sample_images = (None, None, None)
    sample_info = {}

    tbar = tqdm(dataloader, desc=f"Epoch {epoch} [{mode.upper()}]")

    with torch.set_grad_enabled(mode == 'train'):
        for i, (image, labels, name) in enumerate(tbar):

            if mode == 'train':
                total_step += 1

            experiment.set_step(total_step)

            image = image.float().to(device)
            labels = labels.float().to(device)
            image, labels = process_image_patches(image, labels, config.image_size, mode)

            if forward_kwargs is None:
                forward_kwargs = {}
            current_forward_kwargs = forward_kwargs.copy()

            forward_params = inspect.signature(net.forward).parameters
            if 'epoch' in forward_params:
                current_forward_kwargs['epoch'] = epoch

            if mode == 'train':
                if config.warm_up_step > 0 and total_step < config.warm_up_step:
                    lr_scale = total_step / config.warm_up_step
                    new_lr = config.learning_rate * lr_scale
                    for g in optimizer.param_groups:
                        g['lr'] = new_lr

                with torch.cuda.amp.autocast(enabled=config.amp):
                    preds = net(image, **current_forward_kwargs)
                    loss_change, diceloss, foclaloss = criterion(preds, labels)

                cd_loss = loss_change.mean()
                optimizer.zero_grad()
                grad_scaler.scale(cd_loss).backward()
                torch.nn.utils.clip_grad_norm_(net.parameters(), config.max_norm, norm_type=2)
                grad_scaler.step(optimizer)
                grad_scaler.update()
            else:
                preds = net(image)
                loss_change, diceloss, foclaloss = criterion(preds, labels)
                cd_loss = loss_change.mean()

            epoch_loss += cd_loss.item()
            preds_sigmoid = torch.sigmoid(preds)
            batch_metrics = metric_collection.forward(preds_sigmoid.float(), labels.int().unsqueeze(1))
            current_lr = optimizer.param_groups[0]['lr']

            metrics_str = (
                f"{Fore.CYAN}E{epoch}[{mode.upper()}]{Style.RESET_ALL} "
                f"{Fore.RED}Loss:{cd_loss.item():.4f}{Style.RESET_ALL} | "
                f"{Fore.GREEN}Acc:{batch_metrics['accuracy'].item():.3f}{Style.RESET_ALL} | "
                f"{Fore.YELLOW}F1:{batch_metrics['f1score'].item():.3f}{Style.RESET_ALL} | "
                f"{Fore.BLUE}IoU:{batch_metrics['iou'].item():.3f}{Style.RESET_ALL} | "
                f"{Fore.MAGENTA}P:{batch_metrics['precision'].item():.3f}{Style.RESET_ALL} | "
                f"{Fore.MAGENTA}R:{batch_metrics['recall'].item():.3f}{Style.RESET_ALL}"
            )
            if mode == 'train':
                metrics_str += f" | {Fore.CYAN}LR:{current_lr:.1e}{Style.RESET_ALL}"
            tbar.set_description(metrics_str)

            # if i == 0 and image.shape[0] > 0:
            #     sample_index = np.random.randint(low=0, high=image.shape[0])
            #     large_image_name = name[0]
            #     sample_info = {"name": large_image_name, "patch_idx": sample_index}
            #     input_tensor_to_log = de_normalize(image[sample_index])
            #     img_log = to_pilimg(input_tensor_to_log)
            #     lbl_log = to_pilimg(labels[sample_index].cpu().clone().float())
            #     pred_log = to_pilimg(torch.round(preds_sigmoid[sample_index, 0]).cpu().clone().float())
            #     sample_images = (img_log, lbl_log, pred_log)

            del image, labels, preds, preds_sigmoid

    avg_epoch_loss = epoch_loss / len(dataloader) if len(dataloader) > 0 else 0
    epoch_metrics = metric_collection.compute()
    final_epoch_metrics = {k: v.item() for k, v in epoch_metrics.items()}
    final_epoch_metrics['loss_total'] = avg_epoch_loss

    print_epoch_summary(epoch, mode, final_epoch_metrics)
    log_epoch_metrics_to_excel(epoch, mode, final_epoch_metrics)
    _log_epoch_summary_to_comet(experiment, final_epoch_metrics, sample_images, epoch, mode, sample_info)

    if mode == 'val' or mode == 'test':
        paths = (checkpoint_path, best_f1score_model_path, best_loss_model_path)
        best_metrics = _handle_best_model_saving(
            net, optimizer, scheduler, total_step, final_epoch_metrics, epoch, best_metrics, config, paths, mode
        )

    if mode == 'train':
        return experiment, net, optimizer, grad_scaler, total_step, optimizer.param_groups[0]['lr']
    else:
        return experiment, net, optimizer, total_step, optimizer.param_groups[0]['lr'], best_metrics, final_epoch_metrics

